import os
import smtplib
import threading
import time
import tkinter as tk
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from tkinter import filedialog

import webview
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service

# 定义全局变量
is_running = False  # 标记是否正在运行
login_event = threading.Event()  # 登录成功的事件
start_event = threading.Event()  # 登录成功的事件
is_exit = False  # 是否退出程序的标记
g_window = None  # 全局窗口变量，用于与JS通信


def save_emails_to_excel(all_emails):
    """
    将邮箱地址保存到Excel文件中
    :param all_emails: 邮箱地址列表
    """
    wb = Workbook()
    ws = wb.active

    for index, email in enumerate(all_emails, start=1):
        ws.cell(row=index, column=1, value=email)

    current_date = datetime.now().strftime("%Y-%m-%d")
    timestamp = int(time.time())
    filename = f"./all_emails_{current_date}_{timestamp}.xlsx"
    wb.save(filename)
    print(f"邮箱地址已保存到：{filename}")


def send_stop_to_js():
    """
    调用JavaScript中的stop函数，并修改运行状态
    """
    global g_window, is_running
    global is_exit
    global login_event
    global start_event
    if g_window:
        g_window.evaluate_js('stop()')
    is_running = False  # 标记是否正在运行
    login_event = threading.Event()  # 登录成功的事件
    start_event = threading.Event()  # 登录成功的事件
    is_exit = False  # 是否退出程序的标记
    print("已通知停止")


def remove_em_tags(list_of_strings):
    # Use list comprehension and the replace method to remove <em> and </em>
    cleaned_list = [s.replace('<em>', '').replace('</em>', '') for s in list_of_strings]
    return cleaned_list


def driver(search_keyword, start_page):
    """
    使用Selenium驱动浏览器进行数据抓取
    :param search_keyword: 搜索关键词
    :param start_page: 开始的页码
    """
    global is_running, is_exit
    try:
        # 替换下面的路径为您的ChromeDriver路径
        driver_path = './msedgedriver.exe'
        # 创建Service对象并指定ChromeDriver的路径
        service = Service(executable_path=driver_path)
        with webdriver.Edge(service=service) as driver:
            driver.get('https://www.qcc.com/')
            print("等待登录...")
            print(driver)
            login_event.wait()  # 等待登录成功的事件
            print(driver)
            print("登录成功，继续执行")

            driver.get('https://www.qcc.com/')
            time.sleep(1)
            start_event.wait()
            # search_box = driver.find_element("id", 'searchKey')
            # search_box.send_keys(search_keyword + Keys.ENTER)

            all_emails = []
            while not is_exit:
                try:
                    time.sleep(5)
                    jump_input = driver.find_element("css selector", '.input-jump input.form-control')
                    jump_input.clear()  # 先清空输入框
                    jump_input.send_keys(str(start_page) + Keys.ENTER)
                    time.sleep(5)
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, "html.parser")
                    mailto_links = soup.find_all('a', href=lambda href: href and "mailto:" in href)
                    mailto_addresses = [link['href'].replace("mailto:", "") for link in mailto_links]
                    mailto_addresses = remove_em_tags(mailto_addresses)
                    print(f"页数：{start_page}, 邮箱：{mailto_addresses}")
                    start_page = int(start_page) + 1
                    if start_page % 30 == 0:
                        time.sleep(60 * 5)
                    all_emails.extend(mailto_addresses)

                except Exception as e:
                    print(f"爬取过程中发生错误: {e}")
                    break

        save_emails_to_excel(all_emails)
        send_stop_to_js()

    except Exception as e:
        print(f"驱动程序发生错误！")
    finally:
        is_running = False
        send_stop_to_js()


def send_email(sender_email, sender_password, subject_entry, body_text, receiver_email, attachment_files):
    try:
        # SMTP服务器地址和端口
        smtp_server = "smtp.qq.com"
        smtp_port = 465

        # 构造邮件对象
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['Subject'] = subject_entry

        # 邮件正文，设定为HTML格式
        msg.attach(MIMEText(body_text, 'html'))

        # 添加附件
        for file_path in attachment_files:
            part = MIMEBase('application', "octet-stream")
            with open(file_path, 'rb') as file:  # 以二进制读模式打开文件
                part.set_payload(file.read())  # 读取文件内容
            encoders.encode_base64(part)  # Base64编码
            part.add_header('Content-Disposition', f'attachment; filename={file_path}')  # 添加文件头
            msg.attach(part)  # 将附件添加到邮件消息对象中

        # 将邮件对象转换为字符串
        text = msg.as_string()

        # 创建SMTP连接
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(sender_email, sender_password)  # 登录邮箱

        # 发送邮件
        for email_address in receiver_email:
            server.sendmail(sender_email, email_address, text)
            print(email_address + " 邮件发送成功")

        server.quit()  # 断开服务器连接

    except Exception as e:
        print(f"发送邮件到 {receiver_email} 失败: {e}")


class Api:
    def start(self, search_keyword, start_page="1"):
        """
        开始爬虫任务
        :param search_keyword: 搜索关键词
        :param start_page: 起始页码
        """
        global is_running, is_exit
        if is_running:
            return "正在运行中"

        is_running = True
        is_exit = False
        threading.Thread(target=driver, args=(search_keyword, start_page)).start()
        return "已开始"

    def login_success(self):
        """
        登录成功时调用，设置登录事件
        """
        login_event.set()

    def start_success(self):
        """
        登录成功时调用，设置登录事件
        """
        start_event.set()

    def stop(self):
        """
        停止爬虫任务，并尝试保存已爬取的数据
        """
        global is_exit
        is_exit = True
        print("尝试停止爬虫任务")

    def start_send(self, sender_email, sender_password_entry, subject_entry, body_text, attachment_files):
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        root.quit()
        root.destroy()

        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            receiver_email = []
            for cell in ws["A"]:
                receiver_email.append(cell.value)

            send_email(sender_email, sender_password_entry, subject_entry, body_text, receiver_email, attachment_files)

        except Exception as e:
            print("邮件excel解析失败，请查看文件是否正确？")

    def set_file(self):
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        root.quit()
        root.destroy()
        return file_path


if __name__ == '__main__':
    # 初始化API和webview窗口
    api = Api()

    # 构建HTML文件的路径
    html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'views/index/index.html')
    url = 'file://' + html_file_path

    # 创建并配置webview窗口
    g_window = webview.create_window('邮箱自动收集和发送工具', url=url, js_api=api)
    # webview.start(debug=True)
    webview.start()
