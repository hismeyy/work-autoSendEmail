import smtplib
import threading
import time
import tkinter as tk
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from tkinter import scrolledtext, messagebox, filedialog

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from ttkbootstrap import Style
from webdriver_manager.chrome import ChromeDriverManager

# 初始化事件
start_event = threading.Event()
login_event = threading.Event()
is_execution_running = False
is_not_exit = True


def start_execution(keyword_entry, page_entry):
    global is_execution_running
    global is_not_exit
    global start_event

    if is_execution_running:
        messagebox.showwarning("警告", "自动化进程已在运行！")
        return

    search_keyword = keyword_entry.get().strip()
    start_page = page_entry.get().strip()

    if not search_keyword or not start_page.isdigit():
        messagebox.showwarning("警告", "请提供有效的关键字和起始页码！")
        return

    is_execution_running = True
    start_event.set()
    threading.Thread(target=driver, args=(search_keyword, start_page)).start()


def driver(search_keyword, start_page):
    global is_execution_running
    global is_not_exit
    global start_event

    try:
        service = Service(ChromeDriverManager().install())
        with webdriver.Chrome(service=service) as driver:
            driver.get('https://www.qcc.com/')
            login_event.wait()  # 等待登录成功事件被设置
            login_event.clear()
            driver.get('https://www.qcc.com/')
            time.sleep(3)

            search_box = driver.find_element(by='id', value='searchKey')
            search_box.send_keys(search_keyword + Keys.ENTER)

            all_emails = []

            while is_not_exit:
                try:
                    time.sleep(3)
                    jump_input = driver.find_element(by=By.CSS_SELECTOR, value='.input-jump input.form-control')
                    jump_input.send_keys(start_page)
                    jump_input.send_keys(Keys.ENTER)
                    start_page = str(int(start_page))
                    time.sleep(3)
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, "html.parser")
                    mailto_links = soup.find_all('a', href=lambda href: href and "mailto:" in href)
                    mailto_addresses = [link['href'].replace("mailto:", "") for link in mailto_links]
                    print(f"页数：{start_page}")
                    print(F"邮箱：{mailto_addresses}")
                    start_page = int(start_page) + 1
                    all_emails.extend(mailto_addresses)
                except Exception as e:
                    messagebox.showerror("错误", f"没有更多的页数")
                    break

            is_execution_running = False
            start_event.clear()
            is_not_exit = True
            save_emails_to_excel(all_emails)

    except Exception as e:
        print(f"发生错误: {e}")
        messagebox.showerror("错误", f"发生错误: {e}")

    finally:
        is_execution_running = False
        start_event.clear()


def save_emails_to_excel(all_emails):
    wb = Workbook()
    ws = wb.active

    for index, email in enumerate(all_emails, start=1):
        ws.cell(row=index, column=1, value=email)

    current_date = datetime.now().strftime("%Y-%m-%d")
    timestamp = time.time()
    wb.save(f"./all_emails_{current_date}_{timestamp}.xlsx")


def app_exit():
    global is_execution_running
    global is_not_exit
    if is_execution_running:
        is_not_exit = False
    else:
        messagebox.showerror("警告", f"请先开始执行")


def send_email(sender_email, receiver_email, subject, body, sender_password):
    try:
        # 发件人邮箱 sender_email
        # 发件人邮箱授权码，不是邮箱密码 sender_password
        # 收件人邮箱列表 receiver_email

        # SMTP服务器地址和端口，QQ邮箱端口为465
        smtp_server = "smtp.qq.com"
        smtp_port = 465

        # 构造邮件
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['Subject'] = subject

        # 邮件内容
        msg.attach(MIMEText(body, 'plain'))

        # 发送邮件
        text = msg.as_string()

        # 创建SMTP连接
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(sender_email, sender_password)

        for email_address in receiver_email:
            server.sendmail(sender_email, email_address, text)
            print(email_address + "邮件发送成功")
            # 关闭连接
        server.quit()

    except Exception as e:
        print(f"Failed to send email to {receiver_email}: {e}")


def start_send(sender_email_entry, subject_entry, body_text_area, sender_password_entry):
    sender_email = sender_email_entry.get()
    sender_password = sender_password_entry.get()
    subject = subject_entry.get()
    body = body_text_area.get("1.0", tk.END).strip()

    if not sender_email or not subject or not body:
        messagebox.showwarning("错误", "请填写对应的邮件内容")
        return

    filepath = filedialog.askopenfilename()

    if not filepath:
        return

    try:
        wb = load_workbook(filepath)
        ws = wb.active

        receiver_email = []
        for cell in ws["A"]:
            receiver_email.append(cell.value)

        send_email(sender_email, receiver_email, subject, body, sender_password)

    except Exception as e:
        print(f"An error occurred while sending emails: {e}")
        messagebox.showerror("Error", f"An error occurred while sending emails: {e}")


def setup_gui():
    root = tk.Tk()
    root.title("邮箱自动收集和发送工具")
    style = Style(theme='sandstone')

    keyword_label = tk.Label(root, text="将要输入的搜索关键字:")
    keyword_label.pack()
    keyword_entry = tk.Entry(root)
    keyword_entry.pack()

    page_label = tk.Label(root, text="从第几页开始:")
    page_label.pack()
    page_entry = tk.Entry(root)
    page_entry.pack()

    start_button = tk.Button(root, text="开始执行", command=lambda: start_execution(keyword_entry, page_entry))
    start_button.pack(pady=(5, 5))

    login_button = tk.Button(root, text="登录账号成功", command=login_event.set)
    login_button.pack(pady=(5, 5))

    stop_button = tk.Button(root, text="停止执行", command=app_exit)
    stop_button.pack(pady=(5, 5))

    sender_email_label = tk.Label(root, text="请输入发件人邮箱")
    sender_email_label.pack()
    sender_email_entry = tk.Entry(root)
    sender_email_entry.pack()

    sender_password_label = tk.Label(root, text="请输入邮箱授权码")
    sender_password_label.pack()
    sender_password_entry = tk.Entry(root, show="*")
    sender_password_entry.pack()

    subject_label = tk.Label(root, text="请输入邮件主题:")
    subject_label.pack()
    subject_entry = tk.Entry(root)
    subject_entry.pack()

    start_button = tk.Button(root, text="书写邮件", command=lambda: start_execution(keyword_entry, page_entry))
    start_button.pack(pady=(5, 5))

    start_button = tk.Button(root, text="添加附件", command=lambda: start_execution(keyword_entry, page_entry))
    start_button.pack(pady=(5, 5))

    body_label = tk.Label(root, text="请输入邮件内容:")
    body_label.pack()
    body_text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=40, height=10)
    body_text_area.pack()

    # html_label = HTMLLabel(root,
    #                        html='<h1>富文本框示例</h1><p>这是一个包含图片的富文本框</p><img src="example.jpg" alt="Example Image" />')
    # html_label.pack(padx=10, pady=10)

    send_button = tk.Button(root, text="开始发送",
                            command=lambda: start_send(sender_email_entry, subject_entry, body_text_area,
                                                       sender_password_entry))
    send_button.pack(pady=(5, 5))

    root.mainloop()


if __name__ == "__main__":
    setup_gui()
